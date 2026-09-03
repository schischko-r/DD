#!/usr/bin/env python3
"""Build an Excel extract of every unmet DD metric and its available uplift."""

from __future__ import annotations

import json
from collections import defaultdict
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "gravity-app" / "public" / "report-data.json"
OUTPUT = ROOT / "artifacts" / "Аплифт_невыполненных_метрик_DD.xlsx"
HUNDRED = Decimal("100")
ONE_DECIMAL = Decimal("0.1")


def as_decimal(value: object) -> Decimal | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def metric_is_included(metric: dict) -> bool:
    max_value = as_decimal(metric.get("max_value"))
    dd_calculation_flg = as_decimal(metric.get("dd_calculation_flg"))
    return (
        max_value is not None
        and max_value > 0
        and metric.get("is_applicabble_flg") is not False
        and metric.get("excluded_from_index") is not True
        and dd_calculation_flg != 0
    )


def rounded(value: Decimal) -> float:
    return float(value.quantize(ONE_DECIMAL, rounding=ROUND_HALF_UP))


def build_rows(products: list[dict]) -> tuple[list[list[object]], dict[tuple[str, str], Decimal]]:
    rows: list[list[object]] = []
    raw_uplift_by_product: dict[tuple[str, str], Decimal] = defaultdict(Decimal)

    for product in products:
        unit = str(product.get("unit") or "Не указано").strip()
        product_name = str(product.get("name") or "Не указано").strip()
        period = str(product.get("period") or "").strip()
        product_key = (unit, product_name)

        included_metrics: list[tuple[int, int, dict, dict, Decimal, Decimal]] = []
        denominator = Decimal("0")
        for block_order, block in enumerate(product.get("metrics") or []):
            for metric_order, metric in enumerate(block.get("metrics") or []):
                if not metric_is_included(metric):
                    continue
                max_value = as_decimal(metric.get("max_value"))
                value = as_decimal(metric.get("value"))
                if max_value is None or value is None:
                    raise ValueError(
                        f"Missing numeric value for {unit} / {product_name} / "
                        f"{block.get('name')} / {metric.get('name')}"
                    )
                denominator += max_value
                included_metrics.append(
                    (block_order, metric_order, block, metric, value, max_value)
                )

        if denominator <= 0:
            raise ValueError(f"Zero DD denominator for {unit} / {product_name}")

        for block_order, metric_order, block, metric, value, max_value in included_metrics:
            gap = max(max_value - value, Decimal("0"))
            if gap <= 0:
                continue
            uplift_raw = HUNDRED * gap / denominator
            raw_uplift_by_product[product_key] += uplift_raw
            rows.append(
                [
                    unit,
                    product_name,
                    period,
                    str(block.get("name") or "Не указано").strip(),
                    str(metric.get("name") or "Не указано").strip(),
                    str(metric.get("code") or "").strip(),
                    float(value),
                    float(max_value),
                    float(gap),
                    rounded(uplift_raw),
                    block_order,
                    metric_order,
                ]
            )

    rows.sort(key=lambda row: (row[0], row[1], row[10], row[11], row[4]))
    return rows, raw_uplift_by_product


def validate_rows(products: list[dict], rows: list[list[object]], raw_uplift_by_product: dict) -> None:
    if not rows:
        raise ValueError("No unmet DD metrics found")
    if any(row[8] <= 0 or row[9] <= 0 for row in rows):
        raise AssertionError("Every exported row must have a positive gap and uplift")

    expected_by_product: dict[tuple[str, str], Decimal] = {}
    for product in products:
        key = (
            str(product.get("unit") or "Не указано").strip(),
            str(product.get("name") or "Не указано").strip(),
        )
        denominator = Decimal("0")
        earned = Decimal("0")
        for block in product.get("metrics") or []:
            for metric in block.get("metrics") or []:
                if not metric_is_included(metric):
                    continue
                value = as_decimal(metric.get("value"))
                max_value = as_decimal(metric.get("max_value"))
                if value is None or max_value is None:
                    raise ValueError(f"Missing value in validation for {key}")
                denominator += max_value
                earned += min(max(value, Decimal("0")), max_value)
        expected_by_product[key] = HUNDRED * (denominator - earned) / denominator

    for key, expected in expected_by_product.items():
        actual = raw_uplift_by_product.get(key, Decimal("0"))
        if abs(expected - actual) > Decimal("0.0000001"):
            raise AssertionError(f"Product uplift mismatch for {key}: {actual} != {expected}")


def write_workbook(rows: list[list[object]]) -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Невыполненные метрики"

    headers = [
        "Юнит",
        "Продукт",
        "Блок метрик",
        "Метрика",
        "Факт",
        "Макс",
        "Uplift, п.п.",
        "Берут ли в работу",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([row[0], row[1], row[3], row[4], row[6], row[7], row[9], 0])

    header_fill = PatternFill("solid", fgColor="E8F2FF")
    header_font = Font(bold=True, color="1F2D3D")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    uplift_fill = PatternFill("solid", fgColor="E6F4EA")
    uplift_font = Font(bold=True, color="188038")
    for row_index in range(2, sheet.max_row + 1):
        for column in range(5, 7):
            sheet.cell(row_index, column).number_format = "0.###"
        uplift_cell = sheet.cell(row_index, 7)
        uplift_cell.number_format = '0.0 "п.п."'
        uplift_cell.fill = uplift_fill
        uplift_cell.font = uplift_font
        sheet.cell(row_index, 8).number_format = "0"

    widths = {
        "A": 12,
        "B": 28,
        "C": 27,
        "D": 48,
        "E": 14,
        "F": 14,
        "G": 16,
        "H": 20,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:H{sheet.max_row}"
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 32

    table = Table(displayName="UnmetDdMetrics", ref=f"A1:H{sheet.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    workbook.save(OUTPUT)


def validate_workbook(expected_rows: int) -> None:
    workbook = load_workbook(OUTPUT, read_only=False, data_only=True)
    sheet = workbook["Невыполненные метрики"]
    if sheet.max_row != expected_rows + 1 or sheet.max_column != 8:
        raise AssertionError(
            f"Unexpected workbook shape: {sheet.max_row - 1} rows x {sheet.max_column} columns"
        )
    if sheet["C1"].value != "Блок метрик":
        raise AssertionError("Metric block column is missing")
    if sheet["G1"].value != "Uplift, п.п.":
        raise AssertionError("Uplift column is missing")
    if sheet["H1"].value != "Берут ли в работу":
        raise AssertionError("Work flag column is missing")
    if any(sheet.cell(row, 8).value != 0 for row in range(2, sheet.max_row + 1)):
        raise AssertionError("Every work flag must be zero")
    workbook.close()


def main() -> None:
    data = json.loads(SOURCE.read_text(encoding="utf-8"))
    products = data.get("products") or []
    rows, raw_uplift_by_product = build_rows(products)
    validate_rows(products, rows, raw_uplift_by_product)
    write_workbook(rows)
    validate_workbook(len(rows))
    print(f"Saved {OUTPUT.relative_to(ROOT)}: {len(rows)} unmet metrics across {len(products)} products")


if __name__ == "__main__":
    main()
